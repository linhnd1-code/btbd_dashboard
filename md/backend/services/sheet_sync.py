"""
Đồng bộ dữ liệu đội xe trực tiếp từ Google Sheet (không cần service account):
Sheet đã được chia sẻ "Anyone with the link can view" nên có thể tải toàn bộ
workbook (mọi tab) qua URL export XLSX công khai, rồi đọc bằng openpyxl.

Vẫn giữ nguyên cách map cột (theo vị trí) như fleet_parser.py — cấu trúc cột
trong Sheet không đổi, chỉ số dòng dữ liệu tăng theo thời gian.
"""

import io
import time
from datetime import datetime

import openpyxl
import requests
from core.timezone import now_hanoi

SHEET_ID = "1E_fYmNK3TIEMt3xz7JTwG5kmttz_lKbr8xtWRYTDQOY"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

SHEET_MAINTENANCE_LOG = "Lịch sử Bảo dưỡng - sửa chữa"
SHEET_VEHICLES = "Data xe"
SHEET_MAINTENANCE_STATUS = "Data BTBD"

LAST_SYNC_STATUS: dict = {
    "synced_at": None,
    "records": None,
    "vehicles": None,
    "statuses": None,
    "error": None,
}

# Ngưỡng tối thiểu để coi 1 lần tải là "hợp lệ" — thấp hơn NHIỀU so với quy mô thật hiện tại
# (~200 xe, ~3000+ nhật ký) nhưng đủ cao để bắt được trường hợp Google trả về file xlsx
# rút gọn/lỗi khi bị rate-limit hoặc mạng chập chờn (mở được bằng openpyxl, không ném lỗi,
# nhưng chỉ có vài dòng) — xem ghi chú ở sync_from_sheet().
MIN_EXPECTED_VEHICLES = 50
MIN_EXPECTED_RECORDS = 200

# Quan sát thực tế (2026-08-12): Google đôi khi trả về file xlsx export RÚT GỌN (mở được bằng
# openpyxl, không lỗi, nhưng gần như trống) dù request không hề timeout — nghi do sheet quá lớn
# (~14.500 dòng) khiến Google trả bản đang tạo dở/cache lỗi khi bị polling liên tục mỗi 60s.
# Hiện tượng này có vẻ NHẤT THỜI (thử lại sau vài giây thường ra dữ liệu đủ) nên retry vài lần
# trước khi bỏ cuộc, thay vì huỷ đồng bộ ngay ở lần thử đầu.
MAX_FETCH_ATTEMPTS = 3
RETRY_DELAY_SECONDS = 8


def _to_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).replace(",", "").replace(".", "").strip())
    except ValueError:
        return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _to_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def fetch_workbook() -> openpyxl.Workbook:
    # Sheet log đã phình to tới hàng chục nghìn dòng — 30s từng đủ khi sheet còn nhỏ (~3000 dòng)
    # nhưng giờ hay bị "Read timed out" giữa lúc export/tải file lớn hơn. Tăng lên 90s.
    response = requests.get(EXPORT_URL, timeout=90)
    response.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)


def parse_maintenance_records(ws) -> list[dict]:
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 20 or not row[0]:
            continue
        records.append(
            {
                "plate_number": _to_str(row[0]),
                "vehicle_info": _to_str(row[1]),
                "manufacture_year": _to_int(row[2]),
                "odo": _to_int(row[3]),
                "odo_target": _to_int(row[4]),
                "entry_date": _to_datetime(row[5]),
                "work_type": _to_str(row[6]),
                "maintenance_category": _to_str(row[7]),
                "detail": _to_str(row[8]),
                "garage": _to_str(row[9]),
                "expected_finish_date": _to_datetime(row[10]),
                "exit_date": _to_datetime(row[11]),
                "total_hours": _to_float(row[12]),
                "note": _to_str(row[13]),
                "cost": _to_int(row[14]),
                "managing_department": _to_str(row[15]),
                "area": _to_str(row[16]),
                "week": _to_str(row[17]),
                "odo_overdue": _to_int(row[18]),
                "compliance_check": _to_str(row[19]),
            }
        )
    return records


def parse_vehicles(ws) -> list[dict]:
    vehicles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 39 or not row[1]:
            continue
        vehicles.append(
            {
                "plate_number": _to_str(row[1]),
                "status": _to_str(row[0]),
                "load_capacity": _to_str(row[2]),
                "brand": _to_str(row[3]),
                "vehicle_model": _to_str(row[4]),
                "manufacture_year": _to_int(row[5]),
                "manager_unit": _to_str(row[6]),
                "fleet_team": _to_str(row[7]),
                "lifespan_year": _to_str(row[8]),
                "operation_date": _to_datetime(row[9]),
                "years_used": _to_str(row[10]),
                "registration_number": _to_str(row[11]),
                "registration_date": _to_str(row[12]),
                "chassis_number": _to_str(row[13]),
                "engine_number": _to_str(row[14]),
                "inspection_code": _to_str(row[27]),
                "management_number": _to_str(row[28]),
                "receipt_date": _to_str(row[29]),
                "inspection_expiry": _to_str(row[32]),
                "road_fee_expiry": _to_str(row[33]),
                "registration_expiry": _to_str(row[34]),
                "civil_insurance_expiry": _to_str(row[35]),
                "physical_insurance_expiry": _to_str(row[36]),
                "decal_expiry": _to_str(row[37]),
                "odo": _to_int(row[38]),
                "insurance_provider": _to_str(row[39]) if len(row) > 39 else None,
            }
        )
    return vehicles


def parse_maintenance_status(ws) -> list[dict]:
    statuses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 21 or not row[1]:
            continue
        statuses.append(
            {
                "plate_number": _to_str(row[1]),
                "status": _to_str(row[0]),
                "load_capacity": _to_str(row[2]),
                "brand": _to_str(row[3]),
                "vehicle_model": _to_str(row[4]),
                "manager_unit": _to_str(row[5]),
                "manufacture_year": _to_int(row[6]),
                "current_odo": _to_int(row[7]),
                "odo_update_date": _to_datetime(row[8]),
                "last_maintenance_odo": _to_int(row[9]),
                "last_maintenance_date": _to_datetime(row[10]),
                "last_maintenance_odo_target": _to_int(row[11]),
                "next_maintenance_odo": _to_int(row[13]) if len(row) > 13 else None,
                "remaining_odo": _to_int(row[14]) if len(row) > 14 else None,
                "alert_status": _to_str(row[15]) if len(row) > 15 else None,
                "battery_replace_date": _to_str(row[16]) if len(row) > 16 else None,
                "tire_odo": _to_int(row[17]) if len(row) > 17 else None,
                "tire_replace_date": _to_str(row[18]) if len(row) > 18 else None,
                "tire_used_km": _to_int(row[19]) if len(row) > 19 else None,
                "battery_date": _to_str(row[20]) if len(row) > 20 else None,
                "note": _to_str(row[21]) if len(row) > 21 else None,
            }
        )
    return statuses


def sync_from_sheet(db) -> dict:
    """Tải trực tiếp từ Google Sheet (link công khai) và nạp lại toàn bộ 3 bảng dữ liệu."""
    from models.maintenance_record import MaintenanceRecord
    from models.vehicle import Vehicle
    from models.vehicle_maintenance_status import VehicleMaintenanceStatus

    # BẢO VỆ DỮ LIỆU — lỗi thật đã xảy ra (2026-08-12): Google đôi khi trả về file xlsx bị
    # rút gọn/lỗi (nghi do sheet đã phình to ~14.500 dòng, bị polling mỗi 60s). openpyxl vẫn mở
    # được file đó mà KHÔNG ném lỗi — chỉ parse ra rất ít dòng. Hiện tượng này có vẻ NHẤT THỜI
    # (lần gọi tay ngay sau đó thường ra đủ dữ liệu), nên thử lại vài lần trước khi bỏ cuộc, và
    # dù thử lại hết vẫn không đủ thì HUỶ đồng bộ (raise) trước khi đụng tới DB — tuyệt đối không
    # DELETE + INSERT bằng dữ liệu rác, vì 1 lần tải hỏng từng xoá sạch dữ liệu thật đang có.
    records, vehicles, statuses = [], [], []
    failure_reason = None
    for attempt in range(1, MAX_FETCH_ATTEMPTS + 1):
        workbook = fetch_workbook()
        records = parse_maintenance_records(workbook[SHEET_MAINTENANCE_LOG])
        vehicles = parse_vehicles(workbook[SHEET_VEHICLES])
        statuses = parse_maintenance_status(workbook[SHEET_MAINTENANCE_STATUS])

        if (
            len(vehicles) >= MIN_EXPECTED_VEHICLES
            and len(records) >= MIN_EXPECTED_RECORDS
        ):
            failure_reason = None
            break

        failure_reason = f"vehicles={len(vehicles)}, records={len(records)}"
        if attempt < MAX_FETCH_ATTEMPTS:
            print(
                f"[sheet_sync] Lần thử {attempt}/{MAX_FETCH_ATTEMPTS} ra dữ liệu bất thường ít ({failure_reason}) — thử lại sau {RETRY_DELAY_SECONDS}s"
            )
            time.sleep(RETRY_DELAY_SECONDS)

    if failure_reason:
        raise ValueError(
            f"Dữ liệu tải về bất thường ít hơn dự kiến sau {MAX_FETCH_ATTEMPTS} lần thử ({failure_reason}) "
            "— huỷ đồng bộ để tránh xoá nhầm dữ liệu thật đang có trong DB."
        )

    db.query(MaintenanceRecord).delete()
    db.query(Vehicle).delete()
    db.query(VehicleMaintenanceStatus).delete()

    if records:
        db.bulk_insert_mappings(MaintenanceRecord, records)
    if vehicles:
        db.bulk_insert_mappings(Vehicle, vehicles)
    if statuses:
        db.bulk_insert_mappings(VehicleMaintenanceStatus, statuses)
    db.commit()

    result = {
        "records": len(records),
        "vehicles": len(vehicles),
        "statuses": len(statuses),
        "synced_at": now_hanoi().isoformat(),
    }
    LAST_SYNC_STATUS.update(result)
    LAST_SYNC_STATUS["error"] = None
    return result


def record_sync_error(message: str) -> None:
    LAST_SYNC_STATUS["error"] = message
